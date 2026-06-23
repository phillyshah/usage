"""Write the color-coded review workbook (openpyxl).

Sheets:
  * ``Usage``      — the flat, one-row-per-device-line deliverable the accountant
                     works in: ``Source Image Filename`` + the 26 columns from
                     ``reference/output_columns.csv`` (FIELD_GUIDE §5), plus a
                     trailing ``Notes`` review aid.
  * ``Tickets``    — one row per ticket (header reconciliation + round-trip key).
  * ``Line Items`` — one row per device WITH stable Ticket ID / Line ID, the key
                     the corrections re-upload matches on.
  * ``Legend``     — the colour states.

Cell colour follows the confidence model (PROJECT_OVERVIEW principle 2):
    high   -> no fill (confident)
    medium -> amber  FFF2CC (low-confidence guess, eyeball it)
    low    -> red    F4CCCC, cell left BLANK (no confident read, human fills)
    wasted -> yellow FFFF00 on the Price cell (+ WASTED note); still a usage row
Keys, derived counts, and Notes stay uncolored.
"""
from __future__ import annotations

import io
import json
import os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db import db
from app.pipeline.reference import resolve_surgeon

AMBER = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E8EEF1", end_color="E8EEF1", fill_type="solid")
HEADER_FONT = Font(bold=True)

# The 27-column deliverable: Source Image Filename + the 26 output_columns.csv
# columns, in order, plus an appended Notes aid. Each tuple is (header, kind).
# kinds:
#   file        -> uploaded image filename stem (uncolored)
#   blank       -> intentionally blank this version (uncolored)
#   read:<f>    -> ticket header field read off the photo (conf key (None, f))
#   date/month/year -> derived from the ticket surgery_date (conf (None,"surgery_date"))
#   qty         -> always 1 (one row per unit; high)
#   price       -> line unit_price (conf (line,"unit_price")); wasted -> yellow
#   line:<f>    -> line device field (conf (line, f))
#   part:<attr> -> part_info lookup by Ref (conf (line,"description"))
#   surg:<attr> -> surgeon_info lookup by <LastName><DistCode> (surgeon match)
#   notes       -> merged per-line + ticket flags (uncolored)
USAGE_COLUMNS = [
    ("Source Image Filename", "file"),
    ("Reload Code", "blank"),
    ("Surgeon", "read:surgeon"),
    ("DistCode", "read:rep_code"),
    ("Date", "date"),
    ("Month", "month"),
    ("Year", "year"),
    ("Hospital", "surg:hospital"),
    ("Quantity", "qty"),
    ("Price", "price"),
    ("Lot Number", "line:lot"),
    ("Ref Number", "line:ref"),
    ("Expiry Date", "line:expiry_date"),
    ("Invoice No.", "blank"),
    ("Invoice Date", "blank"),
    ("SurgeonName", "surg:surgeon_full_name"),
    ("Distributor", "blank"),
    ("Distributor Rep", "blank"),
    ("Sales Rep", "blank"),
    ("Maxx Sales Manager", "blank"),
    ("Distributing Company", "blank"),
    ("Distributor Code", "surg:dist_code"),
    ("Region", "surg:region"),
    ("Description", "part:description"),
    ("Part Type", "part:part_type"),
    ("Category", "part:category"),
    ("SAP Part Number", "blank"),
    ("Notes", "notes"),  # appended review aid (not in output_columns.csv)
]

# The first 27 headers are the contract (Source Image Filename + the 26 columns).
OUTPUT_CONTRACT_COLUMNS = [h for h, _ in USAGE_COLUMNS[:27]]

TICKET_COLUMNS = [
    ("Ticket ID", None),
    ("Source Image", None),
    ("Entity", "entity"),
    ("Surgery Date", "surgery_date"),
    ("Sales Rep / Distributor", "rep"),
    ("Rep/Distributor Code", "rep_code"),
    ("Surgeon", "surgeon"),
    ("Hospital", "hospital"),
    ("PO Number", "po_number"),
    ("Freight/Delivery Fee", "freight"),
    ("Grand Total", "grand_total"),
    ("Sum of Line Totals", "sum_line_totals"),
    ("Flags / Notes", None),
]

# Raw Extraction: exactly what each source produced per line, BEFORE any lookup
# or resolution. The diagnostic view — shows whether the barcode decoded at all
# and what vision read, next to the final resolved values.
RAW_COLUMNS = [
    "Source Image",
    "Line",
    "Barcode Decoded?",
    "Raw Barcode Payload",
    "Barcode GTIN",
    "Barcode Lot",
    "Barcode Mfg",
    "Barcode Expiry",
    "Barcode Ref (240)",
    "Vision Ref",
    "Vision Lot",
    "Vision Price",
    "Wasted?",
    "Resolved Ref",
    "Resolved Description",
    "Resolved Part Type",
    "Resolved Category",
]

# Per-line sheet WITH the stable IDs the corrections round-trip matches on.
LINE_ITEM_COLUMNS = [
    ("Ticket ID", None),
    ("Line ID", None),
    ("Ref Number", "ref"),
    ("Description", "description"),
    ("Lot Number", "lot"),
    ("Quantity", "qty"),
    ("Mfg Date", "mfg_date"),
    ("Expiry Date", "expiry_date"),
    ("Unit Price", "unit_price"),
    ("Line Total", "line_total"),
    ("Flags / Notes", None),
]


def _parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _date_mdy(v):
    d = _parse_date(v)
    return d.strftime("%m/%d/%Y") if d else None


def _month_num(v):
    d = _parse_date(v)
    return d.month if d else None


def _year_num(v):
    d = _parse_date(v)
    return d.year if d else None


def _file_stem(name):
    """'MO17469-A.jpeg' -> 'MO17469-A'. Returns None for empty."""
    if not name:
        return None
    base = os.path.basename(str(name))
    stem = base.rsplit(".", 1)[0] if "." in base else base
    return stem or None


def _conf_map(ticket_id: str) -> dict:
    """(line_id_or_None, field_name) -> confidence for one ticket."""
    out: dict = {}
    for fe in db.field_extractions_for_ticket(ticket_id):
        out[(fe.get("line_id"), fe.get("field_name"))] = (fe.get("confidence") or "low").lower()
    return out


def _raw_map(ticket_id: str) -> dict:
    """line_id -> raw extraction dict (the JSON snapshot persisted at assembly)."""
    out: dict = {}
    for fe in db.field_extractions_for_ticket(ticket_id):
        if fe.get("field_name") == "raw_blob" and fe.get("line_id"):
            try:
                out[fe["line_id"]] = json.loads(fe.get("orig_value") or "{}")
            except (ValueError, TypeError):
                out[fe["line_id"]] = {}
    return out


def _fill_for(conf: str):
    if conf == "medium":
        return AMBER
    if conf == "low":
        return RED
    return None


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        maxlen = 10
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value is not None:
                    maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(maxlen + 2, 48)


def _flags_text(obj) -> str:
    flags = obj.get("flags") or []
    return "; ".join(flags) if isinstance(flags, list) else str(flags)


def _wasted(line: dict) -> bool:
    flags = line.get("flags") or []
    if isinstance(flags, list):
        return any(str(f).strip().upper() == "WASTED" for f in flags)
    return "WASTED" in str(flags).upper()


def _write_usage_sheet(ws, tickets) -> None:
    """The flat, one-row-per-device-line accountant deliverable."""
    ws.append([h for h, _ in USAGE_COLUMNS])
    _style_header(ws, len(USAGE_COLUMNS))

    for ticket in tickets:
        cmap = _conf_map(ticket["ticket_id"])
        lines = db.lines_for_ticket(ticket["ticket_id"])
        lines.sort(key=lambda x: x.get("created_at") or "")
        ticket_flags = _flags_text(ticket)

        # Surgeon chain resolved once per ticket from the read surgeon + DistCode.
        surg = resolve_surgeon(ticket.get("surgeon"), ticket.get("rep_code"))
        # Hospital cross-check: handwritten hospital vs the looked-up value.
        vis_hosp = ticket.get("hospital")
        hosp_mismatch = bool(
            surg["matched"] and vis_hosp and surg.get("hospital")
            and str(vis_hosp).strip().lower() != str(surg["hospital"]).strip().lower()
        )

        for i, line in enumerate(lines):
            stem = _file_stem(ticket.get("source_filename"))
            line_id = line["line_id"]
            wasted = _wasted(line)
            values: list = []
            fills: list = []

            for header, kind in USAGE_COLUMNS:
                val, conf, override = None, None, None

                if kind == "file":
                    val = stem
                elif kind == "blank":
                    val = None
                elif kind == "notes":
                    parts = [_flags_text(line)]
                    if i == 0 and ticket_flags:
                        parts.append(ticket_flags)
                    val = "; ".join(p for p in parts if p)
                elif kind == "qty":
                    val = line.get("qty") or 1
                elif kind == "price":
                    conf = cmap.get((line_id, "unit_price"), "low")
                    val = None if conf == "low" else line.get("unit_price")
                    if wasted:  # wasted -> yellow regardless of price presence
                        override = YELLOW
                elif kind.startswith("read:"):
                    f = kind.split(":", 1)[1]
                    conf = cmap.get((None, f), "low")
                    val = None if conf == "low" else ticket.get(f)
                elif kind in ("date", "month", "year"):
                    conf = cmap.get((None, "surgery_date"), "low")
                    if conf != "low":
                        sd = ticket.get("surgery_date")
                        val = {"date": _date_mdy, "month": _month_num, "year": _year_num}[kind](sd)
                elif kind.startswith("line:"):
                    f = kind.split(":", 1)[1]
                    conf = cmap.get((line_id, f), "low")
                    val = None if conf == "low" else line.get(f)
                elif kind.startswith("part:"):
                    attr = kind.split(":", 1)[1]
                    # Same provenance as the line's Description confidence.
                    conf = cmap.get((line_id, "description"), "low")
                    pinfo = db.part_info_for_ref(line.get("ref")) if line.get("ref") else None
                    if conf == "low" or not pinfo:
                        conf, val = "low", None
                    else:
                        val = pinfo.get(attr)
                elif kind.startswith("surg:"):
                    attr = kind.split(":", 1)[1]
                    if surg["matched"]:
                        val = surg.get(attr)
                        conf = "high"
                        if attr == "hospital" and hosp_mismatch:
                            conf = "medium"  # cross-check disagreement -> eyeball
                    else:
                        conf, val = "low", None

                values.append(val)
                fills.append(override or _fill_for(conf) if (override or conf) else None)

            ws.append(values)
            r = ws.max_row
            for idx, fill in enumerate(fills, start=1):
                if fill:
                    ws.cell(row=r, column=idx).fill = fill

    _autosize(ws, len(USAGE_COLUMNS))


def _write_tickets_sheet(ws, tickets) -> None:
    ws.append([h for h, _ in TICKET_COLUMNS])
    _style_header(ws, len(TICKET_COLUMNS))
    for ticket in tickets:
        cmap = _conf_map(ticket["ticket_id"])
        row_vals = []
        for header, field in TICKET_COLUMNS:
            if header == "Ticket ID":
                row_vals.append(ticket["ticket_id"])
            elif header == "Source Image":
                row_vals.append(_file_stem(ticket.get("source_filename")) or "")
            elif header == "Flags / Notes":
                row_vals.append(_flags_text(ticket))
            else:
                conf = cmap.get((None, field), "low")
                row_vals.append("" if conf == "low" else ticket.get(field))
        ws.append(row_vals)
        r = ws.max_row
        for idx, (header, field) in enumerate(TICKET_COLUMNS, start=1):
            if field is None:
                continue
            fill = _fill_for(cmap.get((None, field), "low"))
            if fill:
                ws.cell(row=r, column=idx).fill = fill
    _autosize(ws, len(TICKET_COLUMNS))


def _write_line_items_sheet(ws, tickets) -> None:
    ws.append([h for h, _ in LINE_ITEM_COLUMNS])
    _style_header(ws, len(LINE_ITEM_COLUMNS))
    for ticket in tickets:
        cmap = _conf_map(ticket["ticket_id"])
        lines = db.lines_for_ticket(ticket["ticket_id"])
        lines.sort(key=lambda x: x.get("created_at") or "")
        for line in lines:
            line_id = line["line_id"]
            row_vals = []
            for header, field in LINE_ITEM_COLUMNS:
                if header == "Ticket ID":
                    row_vals.append(ticket["ticket_id"])
                elif header == "Line ID":
                    row_vals.append(line_id)
                elif header == "Flags / Notes":
                    row_vals.append(_flags_text(line))
                else:
                    conf = cmap.get((line_id, field), "low")
                    row_vals.append("" if conf == "low" else line.get(field))
            ws.append(row_vals)
            r = ws.max_row
            for idx, (header, field) in enumerate(LINE_ITEM_COLUMNS, start=1):
                if field is None:
                    continue
                fill = _fill_for(cmap.get((line_id, field), "low"))
                if fill:
                    ws.cell(row=r, column=idx).fill = fill
    _autosize(ws, len(LINE_ITEM_COLUMNS))


def _write_raw_extraction_sheet(ws, tickets) -> None:
    """Dump the raw, pre-resolution extraction per line so it's clear exactly
    what the decoder and vision produced (the diagnostic view)."""
    ws.append(RAW_COLUMNS)
    _style_header(ws, len(RAW_COLUMNS))

    for ticket in tickets:
        stem = _file_stem(ticket.get("source_filename"))
        raw_by_line = _raw_map(ticket["ticket_id"])
        lines = db.lines_for_ticket(ticket["ticket_id"])
        lines.sort(key=lambda x: x.get("created_at") or "")
        for i, line in enumerate(lines, start=1):
            raw = raw_by_line.get(line["line_id"], {})
            pinfo = db.part_info_for_ref(line.get("ref")) if line.get("ref") else None
            ws.append([
                stem,
                i,
                "Yes" if raw.get("decoded") else "No",
                raw.get("payload"),
                raw.get("gtin"),
                raw.get("lot"),
                raw.get("mfg"),
                raw.get("expiry"),
                raw.get("ref"),
                raw.get("vis_ref"),
                raw.get("vis_lot"),
                raw.get("vis_price"),
                "Yes" if raw.get("vis_wasted") else "",
                line.get("ref"),
                line.get("description"),
                (pinfo or {}).get("part_type"),
                (pinfo or {}).get("category"),
            ])
            # Flag rows where nothing decoded so they're easy to spot.
            if not raw.get("decoded"):
                ws.cell(row=ws.max_row, column=3).fill = RED
    _autosize(ws, len(RAW_COLUMNS))


def _write_legend_sheet(ws) -> None:
    ws.append(["Color", "Meaning", "What to do"])
    _style_header(ws, 3)
    legend = [
        ("(white / no fill)", "Confident — validated or agreed across sources", "Nothing — trust it", None),
        ("Amber", "Low-confidence guess — single source or a minor disagreement", "Eyeball it; fix if wrong", AMBER),
        ("Red", "Blank / unreadable — no confident read", "Fill it in", RED),
        ("Yellow", "Wasted component (price still counts toward the total)", "Confirm the WASTED note", YELLOW),
    ]
    for color, meaning, todo, fill in legend:
        ws.append([color, meaning, todo])
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill
    ws.append([])
    ws.append(["Note", "Ticket ID and Line ID (Line Items sheet) are stable keys — do not edit them.", ""])
    ws.append(["", "Edit values directly in the colored cells, save, and re-upload.", ""])
    _autosize(ws, 3)


def write_review_workbook(batch_id: str) -> bytes:
    wb = Workbook()
    tickets = db.tickets_for_batch(batch_id)
    tickets.sort(key=lambda t: t.get("created_at") or "")

    ws_u = wb.active
    ws_u.title = "Usage"
    _write_usage_sheet(ws_u, tickets)
    _write_tickets_sheet(wb.create_sheet("Tickets"), tickets)
    _write_line_items_sheet(wb.create_sheet("Line Items"), tickets)
    _write_raw_extraction_sheet(wb.create_sheet("Raw Extraction"), tickets)
    _write_legend_sheet(wb.create_sheet("Legend"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
