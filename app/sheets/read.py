"""Parse a corrected review workbook back into structured rows.

Reads the Tickets and Line Items sheets by header name (robust to column
reordering) and returns the corrected facts keyed by the stable Ticket ID /
Line ID. The learning harvest and the diff both consume this.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

# header text -> our canonical field name
_TICKET_HEADERS = {
    "Ticket ID": "ticket_id",
    "Source Image": "source_image",
    "Entity": "entity",
    "Surgery Date": "surgery_date",
    "Sales Rep / Distributor": "rep",
    "Rep/Distributor Code": "rep_code",
    "Surgeon": "surgeon",
    "Hospital": "hospital",
    "PO Number": "po_number",
    "Freight/Delivery Fee": "freight",
    "Grand Total": "grand_total",
    "Sum of Line Totals": "sum_line_totals",
    "Flags / Notes": "flags",
}
_LINE_HEADERS = {
    "Ticket ID": "ticket_id",
    "Line ID": "line_id",
    "REF (Part No)": "ref",
    "Description": "description",
    "Size": "size",
    "LOT": "lot",
    "Qty": "qty",
    "Mfg Date": "mfg_date",
    "Expiration Date": "expiry_date",
    "Unit Price": "unit_price",
    "Line Total": "line_total",
    "Flags / Notes": "flags",
}


def _header_index(ws, header_map: dict) -> dict:
    """Map canonical field name -> column index (1-based) from row 1."""
    idx: dict = {}
    for col, cell in enumerate(ws[1], start=1):
        label = (str(cell.value).strip() if cell.value is not None else "")
        if label in header_map:
            idx[header_map[label]] = col
    return idx


def _clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def parse_corrected_workbook(data: bytes) -> dict:
    """bytes -> {tickets: {ticket_id: {...header..., lines: {line_id: {...}}}}}.

    Only rows with a Ticket ID are returned. Lines are nested under their ticket.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    tickets: dict = {}

    if "Tickets" in wb.sheetnames:
        ws = wb["Tickets"]
        idx = _header_index(ws, _TICKET_HEADERS)
        tid_col = idx.get("ticket_id")
        if tid_col:
            for row in ws.iter_rows(min_row=2):
                tid = _clean(row[tid_col - 1].value)
                if not tid:
                    continue
                rec = {"ticket_id": tid, "lines": {}}
                for field, col in idx.items():
                    if field == "ticket_id":
                        continue
                    rec[field] = _clean(row[col - 1].value)
                tickets[tid] = rec

    if "Line Items" in wb.sheetnames:
        ws = wb["Line Items"]
        idx = _header_index(ws, _LINE_HEADERS)
        tid_col = idx.get("ticket_id")
        lid_col = idx.get("line_id")
        if tid_col:
            for row in ws.iter_rows(min_row=2):
                tid = _clean(row[tid_col - 1].value)
                if not tid:
                    continue
                lid = _clean(row[lid_col - 1].value) if lid_col else None
                line = {}
                for field, col in idx.items():
                    if field in ("ticket_id",):
                        continue
                    line[field] = _clean(row[col - 1].value)
                tickets.setdefault(tid, {"ticket_id": tid, "lines": {}})
                tickets[tid]["lines"][lid or f"_row{row[0].row}"] = line

    wb.close()
    return {"tickets": tickets}
