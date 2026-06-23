"""Reference masters ingest -> full replace of the product/surgeon crosswalks.

Three read-only sources (FIELD_GUIDE §2), each full-replaced on upload. In
production these are **Excel** workbooks (the same as the Expiry Log); the bundled
repo copies are CSV exports of the same sheets. The loader accepts either format.

  * GTIN_Codes   header row 1: STATUS,GTIN_14,GTIN_12_UPC,PACKAGING_TYPE,
                 PACKAGING_LEVEL,PRODUCT_DESCRIPTION,SKU
  * part_info    row 1 is junk (1,2,3,4); real header is row 2
                 (Part Number,Description,Part Type,Category); data from row 3.
  * surgeon_info header row 1; a *record* is any row with a non-empty DistCode.
                 Address-overflow rows have blank keys -> skip them.

Columns are matched by header NAME (not position), so column reordering between
exports is tolerated. GTIN_14 is guarded against Excel dropping leading zeros.
"""
from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from app.db import db

# Repo-bundled defaults (committed; product/surgeon data, no PHI).
REFERENCE_DIR = Path(__file__).resolve().parents[2] / "reference"


def _cell_str(v) -> str | None:
    """Stringify a cell from CSV or Excel without scientific notation, BOM, or
    trailing '.0' on integer-valued floats (Excel reads numeric columns as float)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        v = int(v) if v.is_integer() else v
        return str(v)
    s = str(v).strip().lstrip("﻿")
    return s or None


def _clean(v) -> str | None:
    return _cell_str(v)


def _pad_gtin(v: str | None) -> str | None:
    """Excel often stores GTIN_14 as a number, dropping leading zeros. A purely
    numeric GTIN shorter than 14 digits is left-padded back to 14."""
    s = _cell_str(v)
    if s and s.isdigit() and len(s) < 14:
        return s.zfill(14)
    return s


def _is_xlsx(data: bytes) -> bool:
    # .xlsx is a zip; legacy .xls (OLE2) starts with the D0 CF magic.
    return data[:2] == b"PK" or data[:4] == b"\xd0\xcf\x11\xe0"


def _records(data: bytes, header_offset: int = 0, sheet: str | None = None) -> list[dict]:
    """Load a table from CSV or Excel into header-keyed row dicts.

    `header_offset` = number of leading rows before the header row (1 for the
    junk row in part_info). For Excel, reads the named sheet or the first one.
    """
    if _is_xlsx(data):
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        grid = list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))

    if len(grid) <= header_offset:
        return []
    header = [_cell_str(c) for c in grid[header_offset]]
    out: list[dict] = []
    for raw in grid[header_offset + 1:]:
        if not raw or all(c in (None, "") for c in raw):
            continue
        rec = {}
        for i, h in enumerate(header):
            if h:
                rec[h] = _cell_str(raw[i]) if i < len(raw) else None
        out.append(rec)
    return out


def _normalize_distcode(code: str | None) -> str | None:
    """'MC - 001' / ' mc-001 ' -> 'MC-001' (collapse spaces, upper)."""
    if not code:
        return None
    return "".join(str(code).split()).upper() or None


def surgeon_key(last_name: str | None, dist_code: str | None) -> str | None:
    """Build the join key: <SurgeonLastName><DistCode>, normalized upper/no-space."""
    ln = _clean(last_name)
    dc = _normalize_distcode(dist_code)
    if not (ln and dc):
        return None
    return ("".join(ln.split()) + dc).upper()


# ---------------------------------------------------------------------------
# Parsers (CSV or Excel bytes -> list[dict] matching the table columns)
# ---------------------------------------------------------------------------
def parse_gtin_codes(data: bytes) -> list[dict]:
    rows: list[dict] = []
    seen: set[str] = set()
    for r in _records(data, header_offset=0):
        gtin = _pad_gtin(r.get("GTIN_14"))
        if not gtin or gtin in seen:
            continue
        seen.add(gtin)
        rows.append({
            "gtin_14": gtin,
            "gtin_12_upc": _clean(r.get("GTIN_12_UPC")),
            "sku": _clean(r.get("SKU")),
            "product_description": _clean(r.get("PRODUCT_DESCRIPTION")),
            "status": _clean(r.get("STATUS")),
            "packaging_type": _clean(r.get("PACKAGING_TYPE")),
            "packaging_level": _clean(r.get("PACKAGING_LEVEL")),
        })
    return rows


def parse_part_info(data: bytes) -> list[dict]:
    """Row 1 junk, row 2 header, data from row 3. Key = Part Number (exact)."""
    rows: list[dict] = []
    seen: set[str] = set()
    for r in _records(data, header_offset=1):
        part = _clean(r.get("Part Number"))
        if not part or part in seen:
            continue
        seen.add(part)
        rows.append({
            "part_number": part,
            "description": _clean(r.get("Description")),
            "part_type": _clean(r.get("Part Type")),
            "category": _clean(r.get("Category")),
        })
    return rows


def parse_surgeon_info(data: bytes) -> list[dict]:
    """A record is any row with a non-empty DistCode; blank-key overflow rows skip."""
    rows: list[dict] = []
    seen: set[str] = set()
    for r in _records(data, header_offset=0):
        dist = _normalize_distcode(r.get("DistCode"))
        last = _clean(r.get("Surgeon Last Name"))
        if not dist:
            continue  # address-overflow continuation row
        key = surgeon_key(last, dist)
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append({
            "surgeon_distcode": key,
            "surgeon_last_name": last,
            "dist_code": dist,
            "status": _clean(r.get("Status")),
            "distributor": _clean(r.get("Distributor")),
            "distributor_rep": _clean(r.get("DistributorRep")),
            "sales_manager": _clean(r.get("Sales Manager")),
            "maxx_ortho_manager": _clean(r.get("Maxx Ortho Manager")),
            "surgeon_full_name": _clean(r.get("Surgeon Full Name")),
            "hospital": _clean(r.get("Hospital")),
            "region": _clean(r.get("Region")),
        })
    return rows


# ---------------------------------------------------------------------------
# Ingest orchestration
# ---------------------------------------------------------------------------
def ingest_masters(gtin: bytes | None = None,
                   part_info: bytes | None = None,
                   surgeon_info: bytes | None = None,
                   as_of: str | None = None) -> dict:
    """Full-replace whichever masters are supplied. Returns per-table row counts.

    `as_of` stamps the ingest with the data's effective date (e.g. the bundled
    monthly snapshot date) instead of the load time; operator uploads leave it
    None so the upload time is used.
    """
    summary = {"gtin_rows": None, "part_rows": None, "surgeon_rows": None}
    if gtin is not None:
        g = parse_gtin_codes(gtin)
        db.replace_reference_gtin(g)
        summary["gtin_rows"] = len(g)
    if part_info is not None:
        p = parse_part_info(part_info)
        db.replace_reference_part_info(p)
        summary["part_rows"] = len(p)
    if surgeon_info is not None:
        s = parse_surgeon_info(surgeon_info)
        db.replace_reference_surgeons(s)
        summary["surgeon_rows"] = len(s)
    row = summary.copy()
    if as_of:
        row["ingested_at"] = as_of
    db.log_masters_ingest(row)
    return summary


def bundled_as_of() -> str | None:
    """The effective date of the committed master snapshot (reference/MASTERS_VERSION).

    Bumped whenever the bundled files are refreshed; returned as an ISO timestamp
    so the freshness banner shows the data date, not the deploy time.
    """
    try:
        d = (REFERENCE_DIR / "MASTERS_VERSION").read_text().strip()
    except FileNotFoundError:
        return None
    return f"{d}T00:00:00+00:00" if d else None


def load_bundled_masters() -> dict:
    """Seed the reference masters from the repo-bundled files (offline/dev/tests).

    Prefers an Excel workbook if present (production format), else the CSV export.
    """
    stems = {"gtin": "GTIN_Codes", "part_info": "part_info", "surgeon_info": "surgeon_info"}
    kwargs: dict = {}
    for key, stem in stems.items():
        for ext in (".xlsx", ".xls", ".csv"):
            p = REFERENCE_DIR / f"{stem}{ext}"
            if p.exists():
                kwargs[key] = p.read_bytes()
                break
    if not kwargs:
        return {"gtin_rows": None, "part_rows": None, "surgeon_rows": None}
    return ingest_masters(as_of=bundled_as_of(), **kwargs)
