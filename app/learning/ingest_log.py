"""Expiry Log ingest -> full replace of the reference tables.

Only the 'Expiry Log History' tab matters (ignore 'Missing' and 'Each Lot Expiry
Update'). Data starts at row 4; row 3 is headers.

Columns:
  A Part No (REF, primary join key)   B Description (size baked in)
  C Lot #                             D Total Qty Released
  E Lot Pallet                        F Expiry Date
  G Notes (ignore)
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime

from openpyxl import load_workbook

from app.db import db

HISTORY_TAB = "Expiry Log History"


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


_SIZE_RE = re.compile(r"size\s*[:#]?\s*(.+)$", re.IGNORECASE)


def split_size(description: str | None) -> str | None:
    if not description:
        return None
    m = _SIZE_RE.search(description)
    if m:
        return m.group(1).strip(" .,") or None
    return None


def ingest_expiry_log(data: bytes) -> dict:
    """Parse + full-replace reference_lots and reference_parts.

    Returns {row_count, unique_parts, unique_lots}.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    if HISTORY_TAB not in wb.sheetnames:
        # Fall back to the first sheet if the tab isn't named as expected.
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[HISTORY_TAB]

    lots: list[dict] = []
    parts: dict[str, dict] = {}
    seen_lots: set[str] = set()

    # Data starts row 4 (row 3 headers). iter from row 4.
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or all(c in (None, "") for c in row):
            continue
        part_no = row[0]
        if part_no in (None, ""):
            continue
        part_no = str(part_no).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else None
        lot = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else None
        total_qty = _to_int(row[3]) if len(row) > 3 else None
        lot_pallet = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "") else None
        expiry = _to_date(row[5]) if len(row) > 5 else None

        lots.append({
            "part_no": part_no,
            "description": description,
            "lot": lot,
            "total_qty_released": total_qty,
            "lot_pallet": lot_pallet,
            "expiry_date": expiry,
        })
        if lot:
            seen_lots.add(lot.upper())

        # Dedup REF -> description/size (first non-empty description wins).
        key = part_no.upper()
        if key not in parts:
            parts[key] = {
                "part_no": part_no,
                "description": description,
                "size": split_size(description),
            }
        elif not parts[key].get("description") and description:
            parts[key]["description"] = description
            parts[key]["size"] = split_size(description)

    wb.close()

    db.replace_reference(lots, list(parts.values()))
    summary = {
        "row_count": len(lots),
        "unique_parts": len(parts),
        "unique_lots": len(seen_lots),
    }
    db.log_ingest(summary.copy())
    return summary
