"""The flat 'Usage' deliverable: exact contract columns, Source Image Filename,
numeric Date/Month/Year, surgeon + part_info joins, and wasted -> yellow."""
import io

from openpyxl import load_workbook

from app.db import db
from app.pipeline.assemble import assemble_and_persist
from app.sheets.write import OUTPUT_CONTRACT_COLUMNS, write_review_workbook

# Source Image Filename + the 26 output_columns.csv columns, in order.
EXPECTED_CONTRACT = [
    "Source Image Filename", "Reload Code", "Surgeon", "DistCode", "Date", "Month",
    "Year", "Hospital", "Quantity", "Price", "Lot Number", "Ref Number",
    "Expiry Date", "Invoice No.", "Invoice Date", "SurgeonName", "Distributor",
    "Distributor Rep", "Sales Rep", "Maxx Sales Manager", "Distributing Company",
    "Distributor Code", "Region", "Description", "Part Type", "Category",
    "SAP Part Number",
]


def _f(value, confidence="high"):
    return {"value": value, "confidence": confidence}


def _empty_label():
    return {"gtin": None, "lot": None, "expiry": None, "mfg": None,
            "serial": None, "raw": None, "decoded": False, "ref": None}


def test_contract_columns_are_exact():
    assert OUTPUT_CONTRACT_COLUMNS == EXPECTED_CONTRACT


def test_usage_row_values_and_joins():
    db.replace_reference_part_info([
        {"part_number": "USG-REF-1", "description": "Widget Deluxe",
         "part_type": "Gadget", "category": "Hardware"},
    ])
    db.replace_reference_surgeons([
        {"surgeon_distcode": "WOODWORTHGR-ME-001", "surgeon_last_name": "Woodworth",
         "dist_code": "GR-ME-001", "status": "Active",
         "surgeon_full_name": "Avery Woodworth", "hospital": "Sierra Medical",
         "region": "West", "distributor_rep": "Pat Rep"},
    ])
    batch = db.create_batch()
    ticket = db.create_ticket({
        "batch_id": batch["id"], "entity": "Maxx Orthopedics",
        "source_filename": "MO083596.jpg", "surgeon": "Woodworth",
        "rep_code": "GR-ME-001", "hospital": "Sierra Medical",
        "surgery_date": "2026-06-01", "status": "pending_review",
    })
    vision = {
        "header": {"surgeon": _f("Woodworth"), "rep_code": _f("GR-ME-001"),
                   "hospital": _f("Sierra Medical"), "surgery_date": _f("2026-06-01")},
        "lines": [{"index": 0, "ref": _f("USG-REF-1"), "lot": _f("L900"),
                   "qty": _f(1), "unit_price": _f(600)}],
        "freight": _f(None, "low"), "grand_total": _f(600),
    }
    assemble_and_persist(ticket, vision, [_empty_label()])

    wb = load_workbook(io.BytesIO(write_review_workbook(batch["id"])))
    assert wb.sheetnames == ["Usage", "Tickets", "Line Items", "Legend"]
    ws = wb["Usage"]
    headers = [c.value for c in ws[1]]
    assert headers[:27] == EXPECTED_CONTRACT
    assert headers[27] == "Notes"

    row = {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(headers)}
    assert row["Source Image Filename"] == "MO083596"   # extension stripped
    assert row["Reload Code"] is None                    # not used
    assert row["Surgeon"] == "Woodworth"
    assert row["DistCode"] == "GR-ME-001"
    assert row["Date"] == "06/01/2026"                   # MM/DD/YYYY
    assert row["Month"] == 6 and row["Year"] == 2026     # numeric
    assert row["Hospital"] == "Sierra Medical"           # surgeon_info lookup
    assert row["SurgeonName"] == "Avery Woodworth"
    assert row["Distributor Code"] == "GR-ME-001"        # canonical lookup
    assert row["Region"] == "West"
    assert row["Quantity"] == 1
    assert row["Price"] == 600
    assert row["Ref Number"] == "USG-REF-1"
    assert row["Lot Number"] == "L900"
    assert row["Description"] == "Widget Deluxe"
    assert row["Part Type"] == "Gadget"
    assert row["Category"] == "Hardware"
    assert row["SAP Part Number"] is None                # deferred


def test_unmatched_surgeon_leaves_lookups_blank_and_red():
    db.replace_reference_surgeons([])  # no surgeon records
    db.replace_reference_part_info([])
    batch = db.create_batch()
    ticket = db.create_ticket({
        "batch_id": batch["id"], "source_filename": "MH99999-A.jpeg",
        "surgeon": "Nobody", "rep_code": "ZZ-999", "surgery_date": "2026-06-01",
        "status": "pending_review",
    })
    vision = {
        "header": {"surgeon": _f("Nobody"), "rep_code": _f("ZZ-999"),
                   "surgery_date": _f("2026-06-01")},
        "lines": [{"index": 0, "ref": _f(None, "low"), "qty": _f(1),
                   "unit_price": _f(None, "low")}],
        "freight": _f(None, "low"), "grand_total": _f(None, "low"),
    }
    assemble_and_persist(ticket, vision, [_empty_label()])

    wb = load_workbook(io.BytesIO(write_review_workbook(batch["id"])))
    ws = wb["Usage"]
    headers = [c.value for c in ws[1]]
    hcol = headers.index("Hospital") + 1
    cell = ws.cell(row=2, column=hcol)
    assert cell.value in (None, "")                       # no surgeon match -> blank
    assert cell.fill.fgColor.rgb.endswith("F4CCCC")       # red


def test_wasted_price_cell_is_yellow():
    db.replace_reference_part_info([
        {"part_number": "Y-REF", "description": "Y", "part_type": "T", "category": "C"},
    ])
    batch = db.create_batch()
    ticket = db.create_ticket({"batch_id": batch["id"], "source_filename": "MO1.jpg",
                               "surgery_date": "2026-06-01", "status": "pending_review"})
    vision = {
        "header": {"surgery_date": _f("2026-06-01")},
        "lines": [{"index": 0, "ref": _f("Y-REF"), "lot": _f("L1"), "qty": _f(1),
                   "unit_price": _f(150), "wasted": _f(True)}],
        "freight": _f(None, "low"), "grand_total": _f(150),
    }
    assemble_and_persist(ticket, vision, [_empty_label()])

    wb = load_workbook(io.BytesIO(write_review_workbook(batch["id"])))
    ws = wb["Usage"]
    headers = [c.value for c in ws[1]]
    pcol = headers.index("Price") + 1
    assert ws.cell(row=2, column=pcol).fill.fgColor.rgb.endswith("FFFF00")  # yellow
    notes = ws.cell(row=2, column=headers.index("Notes") + 1).value
    assert "WASTED" in (notes or "")
