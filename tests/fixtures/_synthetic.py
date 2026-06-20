"""Generate synthetic fixtures (a ticket image with a real GS1 DataMatrix, and a
small Expiry Log workbook) so the end-to-end path is testable without the real
sample photos. Drop the real fixtures in alongside these for production testing.
"""
from __future__ import annotations

import io

import numpy as np
from openpyxl import Workbook
from PIL import Image
from pylibdmtx.pylibdmtx import encode as dmtx_encode

GS = "\x1d"  # FNC1 / group separator


def gtin14(base13: str) -> str:
    """Append a valid GS1 check digit to a 13-digit base -> 14-digit GTIN."""
    assert len(base13) == 13 and base13.isdigit()
    digits = [int(c) for c in base13]
    # GS1 check digit: from rightmost, alternate x3/x1 (rightmost data digit x3).
    total = 0
    for i, d in enumerate(reversed(digits)):
        total += d * (3 if i % 2 == 0 else 1)
    check = (10 - (total % 10)) % 10
    return base13 + str(check)


def gs1_payload(gtin: str, expiry_yymmdd: str, lot: str) -> str:
    # 01 (GTIN-14, fixed) + 17 (expiry, fixed 6) + 10 (lot, variable, last).
    return f"01{gtin}17{expiry_yymmdd}10{lot}"


def make_ticket_image(payload: str, width: int = 1000, height: int = 1400) -> bytes:
    """White ticket with a DataMatrix placed in the label-grid region."""
    canvas = Image.new("RGB", (width, height), "white")

    encoded = dmtx_encode(payload.encode("ascii"))
    dm = Image.frombytes("RGB", (encoded.width, encoded.height), encoded.pixels)
    # Scale up so the modules are big enough to decode reliably from the photo.
    scale = 4
    dm = dm.resize((encoded.width * scale, encoded.height * scale), Image.NEAREST)

    # Place in the grid region (top-left of the lower 70%).
    canvas.paste(dm, (60, int(height * 0.30)))
    buf = io.BytesIO()
    canvas.save(buf, format="JPEG", quality=95)
    return buf.getvalue()


def make_expiry_log(parts: list[dict]) -> bytes:
    """parts: [{part_no, description, lot, expiry}] -> Expiry Log workbook bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expiry Log History"
    # Rows 1-2 spacer/title, row 3 headers, data from row 4 (matches the spec).
    ws.append(["Expiry Log"])  # row 1
    ws.append([])  # row 2
    ws.append(["Part No", "Description", "Lot #", "Total Qty Released",
               "Lot Pallet", "Expiry Date", "Notes"])  # row 3
    for p in parts:
        ws.append([p["part_no"], p["description"], p["lot"], 10, "P1", p["expiry"], ""])
    # Add the ignored tabs to prove we only read the right one.
    wb.create_sheet("Missing")
    wb.create_sheet("Each Lot Expiry Update")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
