"""Handwritten quantity > 1: _qty parsing + end-to-end through assemble_and_persist."""
import io

import pytest
from openpyxl import load_workbook

from app.db import db
from app.learning.ingest_reference import load_bundled_masters
from app.pipeline.assemble import _qty, assemble_and_persist
from app.sheets.write import write_review_workbook


@pytest.mark.parametrize("raw,expected", [
    ("4", 4),
    ("x4", 4),
    ("Qty 4", 4),
    (4, 4),
    (4.0, 4),
    (None, 1),
    ("", 1),
    ("0", 1),
    (-2, 1),
    ("two", 1),
    (True, 1),
])
def test_qty_parsing(raw, expected):
    assert _qty(raw) == expected


def _f(v, c="high"):
    return {"value": v, "confidence": c}


def _label():
    return {"gtin": "00810008120088", "lot": "S41122707", "expiry": "2028-10-31",
            "mfg": "2023-11-01", "ref": "MO-MSFC-56/MH", "serial": None,
            "raw": "x", "decoded": True}


def _ticket():
    load_bundled_masters()
    b = db.create_batch()
    return b, db.create_ticket({"batch_id": b["id"], "entity": "Maxx Health",
                                "source_filename": "MH.jpg", "status": "pending_review"})


def test_handwritten_qty_drives_line_total():
    """qty=4 @ 25.0 -> line_total 100.0, reconciles with grand_total, no flag."""
    _b, t = _ticket()
    vision = {
        "header": {}, "grand_total": _f(100.0),
        "lines": [{"qty": _f(4), "unit_price": _f(25.0)}],
    }
    assemble_and_persist(t, vision, [_label()])
    lines = db.lines_for_ticket(t["ticket_id"])
    assert len(lines) == 1
    assert lines[0]["qty"] == 4
    assert lines[0]["line_total"] == 100.0
    ticket = db.get_ticket(t["ticket_id"])
    assert not any("Grand total" in f for f in (ticket.get("flags") or []))


def test_missing_qty_defaults_to_one():
    """A line with no qty field persists qty == 1."""
    _b, t = _ticket()
    vision = {
        "header": {}, "grand_total": _f(25.0),
        "lines": [{"unit_price": _f(25.0)}],
    }
    assemble_and_persist(t, vision, [_label()])
    lines = db.lines_for_ticket(t["ticket_id"])
    assert len(lines) == 1
    assert lines[0]["qty"] == 1


def test_workbook_usage_quantity_cell_reflects_qty():
    """The Usage 'Quantity' cell shows the handwritten 4."""
    b, t = _ticket()
    vision = {
        "header": {}, "grand_total": _f(100.0),
        "lines": [{"qty": _f(4), "unit_price": _f(25.0)}],
    }
    assemble_and_persist(t, vision, [_label()])

    wb = load_workbook(io.BytesIO(write_review_workbook(b["id"])))
    ws = wb["Usage"]
    headers = [c.value for c in ws[1]]
    qcol = headers.index("Quantity") + 1
    assert ws.cell(row=2, column=qcol).value == 4
