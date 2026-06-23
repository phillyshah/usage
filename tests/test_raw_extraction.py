"""The Raw Extraction sheet: the diagnostic view of exactly what each source
produced per line, before any lookup/resolution."""
import io

import pytest
from openpyxl import load_workbook

from app.db import db
from app.learning.ingest_reference import load_bundled_masters
from app.pipeline.assemble import assemble_and_persist
from app.sheets.write import RAW_COLUMNS, write_review_workbook


@pytest.fixture(autouse=True)
def _seed():
    load_bundled_masters()


def _f(v, c="high"):
    return {"value": v, "confidence": c}


def _raw_sheet(batch_id):
    wb = load_workbook(io.BytesIO(write_review_workbook(batch_id)), data_only=True)
    assert "Raw Extraction" in wb.sheetnames
    ws = wb["Raw Extraction"]
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


def test_decoded_barcode_shows_raw_payload_and_fields():
    batch = db.create_batch()
    ticket = db.create_ticket({"batch_id": batch["id"], "entity": "Maxx Health",
                               "source_filename": "MH-1.jpg", "status": "pending_review"})
    label = {"gtin": "00810008120088", "lot": "S41122707", "mfg": "2023-11-01",
             "expiry": "2028-10-31", "ref": "MO-MSFC-56/MH", "serial": None,
             "raw": "0100810008120088...", "decoded": True}
    vision = {"header": {}, "lines": [{"unit_price": _f(900.0)}]}
    assemble_and_persist(ticket, vision, [label])

    header, rows = _raw_sheet(batch["id"])
    assert header == tuple(RAW_COLUMNS)
    assert len(rows) == 1
    r = dict(zip(RAW_COLUMNS, rows[0]))
    assert r["Source Image"] == "MH-1"
    assert r["Barcode Decoded?"] == "Yes"
    assert r["Raw Barcode Payload"] == "0100810008120088..."
    assert r["Barcode GTIN"] == "00810008120088"
    assert r["Barcode Lot"] == "S41122707"
    assert r["Barcode Ref (240)"] == "MO-MSFC-56/MH"
    assert r["Vision Price"] == 900.0
    # Resolution still happened: the GTIN master filled the resolved columns.
    assert r["Resolved Ref"] == "MO-MSFC-56/MH"
    assert r["Resolved Part Type"] and r["Resolved Category"]


def test_undecoded_line_marked_no_but_shows_vision():
    """When the barcode fails to decode (an empty padded label) the row is still
    written, marked 'No', with the vision read visible — the exact diagnostic
    for an otherwise-blank output."""
    batch = db.create_batch()
    ticket = db.create_ticket({"batch_id": batch["id"], "entity": "Maxx Health",
                               "source_filename": "MH-2.jpg", "status": "pending_review"})
    empty_label = {"gtin": None, "lot": None, "expiry": None, "mfg": None,
                   "serial": None, "raw": None, "decoded": False, "ref": None}
    vision = {"header": {}, "lines": [{"ref": _f("MO-STVC-35/03"), "unit_price": _f(1900.0)}]}
    assemble_and_persist(ticket, vision, [empty_label])

    _, rows = _raw_sheet(batch["id"])
    r = dict(zip(RAW_COLUMNS, rows[0]))
    assert r["Barcode Decoded?"] == "No"
    assert r["Raw Barcode Payload"] in (None, "")
    assert r["Barcode GTIN"] in (None, "")
    # Vision still read the printed REF/price — shows where the data came from.
    assert r["Vision Ref"] == "MO-STVC-35/03"
    assert r["Vision Price"] == 1900.0
