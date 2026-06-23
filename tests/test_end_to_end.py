"""End-to-end offline test of the full pipeline via the FastAPI app.

Exercises: reference-log ingest -> image upload (redaction gate) -> batch run
(barcode decode + log resolve + assemble + colored workbook) -> corrections
re-upload (harvest + diff + verify).
"""
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.db import db
from app.main import app
from tests.fixtures._synthetic import (
    gs1_payload,
    gtin14,
    make_expiry_log,
    make_ticket_image,
)

client = TestClient(app)

REF = "RAUUX400-RK"
LOT = "LOT12345"
DESC = "Acetabular Shell, Size 54"
GTIN = gtin14("0361414000123")
EXPIRY_YYMMDD = "271130"
EXPIRY_ISO = "2027-11-30"


def test_health():
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json() == {"status": "ok"}


def test_full_pipeline():
    # 1. Ingest the reference log (full replace).
    log_bytes = make_expiry_log([
        {"part_no": REF, "description": DESC, "lot": LOT, "expiry": EXPIRY_ISO},
    ])
    r = client.post(
        "/reference/log",
        files={"file": ("Expiry_Log.xlsx", log_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    summary = r.json()
    assert summary["unique_parts"] == 1
    assert summary["unique_lots"] == 1
    # Seed the part_info master so the lot-recovered REF gets a description.
    db.replace_reference_part_info([
        {"part_number": REF, "description": DESC, "part_type": "Shell",
         "category": "Acetabular Cup"},
    ])

    # 2. Upload a ticket image carrying that lot's DataMatrix.
    payload = gs1_payload(GTIN, EXPIRY_YYMMDD, LOT)
    img = make_ticket_image(payload)
    r = client.post(
        "/images",
        files={"files": ("ticket_ortho_01.jpg", img, "image/jpeg")},
    )
    assert r.status_code == 202, r.text
    body = r.json()
    batch_id = body["batch_id"]
    assert len(body["tickets"]) == 1
    ticket = body["tickets"][0]
    # Redaction must succeed on a known template -> pending_review (not manual).
    assert ticket["status"] == "pending_review", body

    # 3. Run the batch -> colored workbook.
    r = client.post("/batches/run", json={"batch_id": batch_id})
    assert r.status_code == 200, r.text
    run = r.json()
    assert run["ticket_count"] == 1

    # 4. Download + inspect the workbook.
    r = client.get(f"/batches/{batch_id}/sheet")
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Usage", "Tickets", "Line Items", "Legend"]

    ws_lines = wb["Usage"]
    headers = [c.value for c in ws_lines[1]]
    assert "Ref Number" in headers
    # The barcode lot recovered the REF from the Expiry Log.
    rows = list(ws_lines.iter_rows(min_row=2, values_only=True))
    assert rows, "expected at least one line item"
    ref_col = headers.index("Ref Number")
    lot_col = headers.index("Lot Number")
    assert rows[0][ref_col] == REF
    assert rows[0][lot_col] == LOT

    # Price had no vision read -> blank + red fill (low confidence).
    price_col = headers.index("Price") + 1  # openpyxl 1-based
    price_cell = ws_lines.cell(row=2, column=price_col)
    assert price_cell.value in (None, ""), "price should be blank without a vision read"
    assert price_cell.fill.fgColor.rgb.endswith("F4CCCC"), "blank price should be red"

    # REF recovered via the lot (not the GTIN master) -> medium/amber: eyeball it.
    ref_cell = ws_lines.cell(row=2, column=ref_col + 1)
    assert ref_cell.fill.fgColor.rgb.endswith("FFF2CC"), "lot-recovered REF should be amber"

    # 5. Re-upload a corrected sheet (edit the Line Items sheet that carries IDs).
    corrected = _corrected_copy(r_content=r.content, price=750.0)
    r2 = client.post(
        "/corrections/upload",
        files={"files": ("review_corrected.xlsx", corrected,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r2.status_code == 200, r2.text
    res = r2.json()
    assert res["tickets_matched"] == 1
    assert res["tickets_unknown"] == 0


def _corrected_copy(r_content: bytes, price: float) -> bytes:
    """Open the generated workbook, fill the blank Unit Price on the Line Items
    sheet (the one carrying the Ticket/Line IDs the round-trip matches on)."""
    wb = load_workbook(io.BytesIO(r_content))
    ws = wb["Line Items"]
    headers = [c.value for c in ws[1]]
    price_col = headers.index("Unit Price") + 1
    ws.cell(row=2, column=price_col).value = price
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_unknown_template_goes_to_manual_queue():
    # An undecodable/blank image with no template hint cannot be safely redacted
    # against a known template via filename, but defaults to ortho geometry.
    # A truly unreadable (empty) image must NOT be sent anywhere -> manual_queue.
    r = client.post(
        "/images",
        files={"files": ("garbage.jpg", b"not-an-image", "image/jpeg")},
    )
    assert r.status_code == 202
    assert r.json()["tickets"][0]["status"] == "manual_queue"
