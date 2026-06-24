"""Spreadsheet date fields render as MM/DD/YYYY (US), never raw ISO YYYY-MM-DD.

Seeds a processed ticket with ISO date inputs, generates the review workbook, and
asserts the date-valued cells across the sheets come out MM/DD/YYYY. Also unit-tests
the formatting helpers (``_date_mdy`` / ``_fmt_field``) directly.
"""
import io
import re

from openpyxl import load_workbook

from app.db import db
from app.learning.ingest_reference import load_bundled_masters
from app.pipeline.assemble import assemble_and_persist
from app.sheets import write
from app.sheets.write import write_review_workbook

MDY = re.compile(r"^\d{2}/\d{2}/\d{4}$")


def _f(v, c="high"):
    return {"value": v, "confidence": c}


def _labels():
    return [
        {"gtin": "00810008120088", "lot": "S41122707", "expiry": "2028-10-31",
         "mfg": "2023-11-01", "ref": "MO-MSFC-56/MH", "serial": None,
         "raw": "x", "decoded": True},
    ]


def _vision():
    return {
        "header": {"surgery_date": _f("2026-06-15", "high")},
        "grand_total": _f(900),
        "lines": [{"unit_price": _f(900, "high")}],
    }


def _seed():
    load_bundled_masters()
    batch = db.create_batch()
    ticket = db.create_ticket({
        "batch_id": batch["id"], "entity": "Maxx Health",
        "source_filename": "MH.jpg", "surgery_date": "2026-06-15",
        "status": "pending_review"})
    assemble_and_persist(ticket, _vision(), _labels())
    return batch["id"]


def _load(batch_id):
    data = write_review_workbook(batch_id)
    return load_workbook(io.BytesIO(data), data_only=True)


def _cell(ws, header):
    """Index (1-based) of a header in row 1, or raises if missing."""
    headers = [c.value for c in ws[1]]
    return headers.index(header) + 1


def _row_val(ws, header, row=2):
    return ws.cell(row=row, column=_cell(ws, header)).value


def test_line_items_dates_are_mdy():
    wb = _load(_seed())
    ws = wb["Line Items"]
    expiry = _row_val(ws, "Expiry Date")
    mfg = _row_val(ws, "Mfg Date")
    assert expiry == "10/31/2028"
    assert mfg == "11/01/2023"
    assert MDY.match(expiry) and MDY.match(mfg)
    assert "-" not in expiry and "-" not in mfg


def test_raw_extraction_dates_are_mdy():
    wb = _load(_seed())
    ws = wb["Raw Extraction"]
    expiry = _row_val(ws, "Barcode Expiry")
    mfg = _row_val(ws, "Barcode Mfg")
    assert expiry == "10/31/2028"
    assert mfg == "11/01/2023"
    assert "-" not in str(expiry) and "-" not in str(mfg)


def test_tickets_surgery_date_is_mdy():
    wb = _load(_seed())
    ws = wb["Tickets"]
    val = _row_val(ws, "Surgery Date")
    # High-confidence vision read should surface the value as MM/DD/YYYY.
    assert val == "06/15/2026", (
        f"expected surgery date 06/15/2026, got {val!r} "
        "(if empty, this is the confidence-scoring behavior, not a format bug)")
    assert "-" not in str(val)


def test_usage_expiry_date_is_mdy_or_blank():
    wb = _load(_seed())
    ws = wb["Usage"]
    val = _row_val(ws, "Expiry Date")
    # Either blank (low confidence) or MM/DD/YYYY — never raw ISO.
    assert val == "" or val is None or MDY.match(str(val))
    assert "-" not in str(val or "")


def test_no_iso_dates_leak_into_workbook():
    """Sanity: none of the date cells carry a leftover ISO (contains '-')."""
    wb = _load(_seed())
    checks = [
        (wb["Line Items"], "Expiry Date"),
        (wb["Line Items"], "Mfg Date"),
        (wb["Raw Extraction"], "Barcode Expiry"),
        (wb["Raw Extraction"], "Barcode Mfg"),
        (wb["Tickets"], "Surgery Date"),
        (wb["Usage"], "Expiry Date"),
    ]
    for ws, header in checks:
        val = _row_val(ws, header)
        assert "-" not in str(val or ""), f"{ws.title}/{header} leaked ISO: {val!r}"


# ---------------------------------------------------------------------------
# Focused unit tests of the formatting helpers
# ---------------------------------------------------------------------------
def test_date_mdy_helper():
    assert write._date_mdy("2028-10-31") == "10/31/2028"
    assert write._date_mdy("") is None
    assert write._date_mdy(None) is None
    # Idempotent on an already-MDY string.
    assert write._date_mdy("06/15/2026") == "06/15/2026"


def test_fmt_field_passthrough_and_dates():
    assert write._fmt_field("lot", "ABC") == "ABC"
    assert write._fmt_field("expiry_date", "2028-10-31") == "10/31/2028"
