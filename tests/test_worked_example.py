"""End-to-end regression on the §7 worked example (ticket MH13366-A).

Drives the deterministic chain from the six decoded barcode payloads through the
real bundled reference masters: GTIN -> SKU -> part_info, surgeon key -> hospital,
prices injected as the vision read. The six lines must reconcile to $4,136.00 and
the 27-column deliverable must carry the looked-up device + surgeon attributes.
"""
import io

import pytest
from openpyxl import load_workbook

from app.db import db
from app.learning.ingest_reference import load_bundled_masters
from app.pipeline.assemble import assemble_and_persist
from app.pipeline.barcode import decode_single
from app.pipeline.reference import resolve_part, resolve_surgeon
from app.sheets.write import OUTPUT_CONTRACT_COLUMNS, write_review_workbook
from tests.fixtures._worked_example import (
    EXPECTED_HOSPITAL,
    EXPECTED_LINES,
    EXPECTED_REGION,
    GRAND_TOTAL,
    HEADER,
    LABEL_PAYLOADS,
    price_by_ref,
)

@pytest.fixture(autouse=True)
def _seed_masters():
    # The local store is shared across the suite and other tests wipe the
    # masters; re-seed the real bundled CSVs before every test here.
    load_bundled_masters()


def _f(value, confidence="high"):
    return {"value": value, "confidence": confidence}


def test_barcodes_resolve_to_skus_via_gtin_master():
    """Each decoded GTIN resolves to its SKU and part_info description."""
    for exp in EXPECTED_LINES:
        label = decode_single(next(p for p in LABEL_PAYLOADS if p.endswith(exp["ref"])))
        assert label["lot"] == exp["lot"]
        assert label["expiry"] == exp["expiry"]
        part = resolve_part(label.get("ref"), label.get("gtin"), label.get("lot"))
        assert part["ref"] == exp["ref"], exp["ref"]
        assert part["ref_source"] == "gtin"          # SKU from the GTIN master
        assert part["in_gtin_master"] is True
        assert part["in_part_info"] is True
        assert part["description"]                    # part_info filled it
        assert part["part_type"] and part["category"]


def test_surgeon_key_resolves():
    surg = resolve_surgeon(HEADER["surgeon"], HEADER["rep_code"])
    assert surg["matched"] is True
    assert surg["hospital"] == EXPECTED_HOSPITAL
    assert surg["region"] == EXPECTED_REGION
    assert surg["dist_code"] == "MC-001"


def _build_ticket():
    batch = db.create_batch()
    ticket = db.create_ticket({
        "batch_id": batch["id"],
        "entity": HEADER["entity"],
        "surgeon": HEADER["surgeon"],
        "rep_code": HEADER["rep_code"],
        "surgery_date": HEADER["surgery_date"],
        "hospital": HEADER["hospital"],
        "source_filename": "MH13366-A.jpeg",
        "status": "pending_review",
    })
    labels = [decode_single(p) for p in LABEL_PAYLOADS]
    vlines = [
        {"index": i, "qty": _f(1), "unit_price": _f(price_by_ref(lbl["ref"]))}
        for i, lbl in enumerate(labels)
    ]
    vision = {
        "header": {k: _f(v) for k, v in HEADER.items()},
        "lines": vlines,
        "freight": _f(0),
        "grand_total": _f(GRAND_TOTAL),
    }
    assemble_and_persist(ticket, vision, labels)
    return batch, ticket


def test_six_lines_reconcile_to_grand_total():
    batch, ticket = _build_ticket()
    lines = db.lines_for_ticket(ticket["ticket_id"])
    assert len(lines) == 6
    total = sum(ln["line_total"] for ln in lines)
    assert round(total, 2) == GRAND_TOTAL
    refs = {ln["ref"] for ln in lines}
    assert refs == {e["ref"] for e in EXPECTED_LINES}
    # No reconciliation flag — sum matches the grand total.
    fresh = db.get_ticket(ticket["ticket_id"])
    assert not any("Grand total" in f for f in (fresh.get("flags") or []))


def test_usage_sheet_has_contract_columns_and_joined_values():
    batch, ticket = _build_ticket()
    wb = load_workbook(io.BytesIO(write_review_workbook(batch["id"])))
    assert wb.sheetnames == ["Usage", "Tickets", "Line Items", "Raw Extraction", "Legend"]

    ws = wb["Usage"]
    headers = [c.value for c in ws[1]]
    # First 27 columns are the exact contract (Source Image Filename + 26).
    assert headers[:27] == OUTPUT_CONTRACT_COLUMNS
    assert headers[27] == "Notes"

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 6
    col = {h: i for i, h in enumerate(headers)}
    for r in rows:
        assert r[col["Source Image Filename"]] == "MH13366-A"
        assert r[col["Reload Code"]] is None            # not used
        assert r[col["Quantity"]] == 1                  # one row per unit
        assert r[col["Hospital"]] == EXPECTED_HOSPITAL  # surgeon_info lookup
        assert r[col["Region"]] == EXPECTED_REGION
        assert r[col["Surgeon"]] == HEADER["surgeon"]
        assert r[col["Month"]] == 6 and r[col["Year"]] == 2026
        assert r[col["Ref Number"]] in {e["ref"] for e in EXPECTED_LINES}
        assert r[col["Description"]]                     # part_info lookup
        assert r[col["Part Type"]] and r[col["Category"]]
        assert r[col["SAP Part Number"]] is None        # deferred/blank
